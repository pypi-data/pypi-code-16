"""
test_group_wb
----------------------------------

Tests for the `kifield.group_wb` function
"""

import unittest
import hypothesis
import hypothesis.strategies as st
import openpyxl as pyxl

from kifield import kifield

class TestGroupWb(unittest.TestCase):
    def test_groups(self):
        wb = pyxl.Workbook()
        ws = wb.active
        header = ('Ref', 'x', 'y', 'z')
        ws.append(header)
        ws.append(('C1', '1', '1', '1'))
        ws.append(('C2', '1', '1', '1'))
        ws.append(('C3', '1', '1', '1'))

        wb = kifield.group_wb(wb)
        ws = wb.active

        assert ws.max_row == 2
        assert ws.max_column == 4

        values = tuple(ws.values)
        assert values[0] == header
        assert values[1] == ('C1-C3', '1', '1', '1')

    def test_groups2(self):
        wb = pyxl.Workbook()
        ws = wb.active
        header = ('Ref', 'x', 'y', 'z')
        ws.append(header)
        ws.append(('C1', '1', '1', '1'))
        ws.append(('R3', '2', '1', '1'))
        ws.append(('R5', '2', '1', '1'))
        ws.append(('X1', '3', '1', '1'))
        ws.append(('X2', '1', '3', '1'))
        ws.append(('X3', '1', '1', '3'))

        wb = kifield.group_wb(wb)
        ws = wb.active

        assert ws.max_row == 6
        assert ws.max_column == 4

        values = tuple(ws.values)
        assert values[0] == header
        assert values[1] == ('C1', '1', '1', '1')
        assert values[2] == ('R3, R5', '2', '1', '1')
        assert values[3] == ('X1', '3', '1', '1')
        assert values[4] == ('X2', '1', '3', '1')
        assert values[5] == ('X3', '1', '1', '3')
